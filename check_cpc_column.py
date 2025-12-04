#!/usr/bin/env python3
"""
Verifica los valores de la columna CPC (columna I)
"""

import openpyxl

xlsx_path = r"c:\Users\User\Downloads\Nueva carpeta (2)\Nueva carpeta (2)\pdftoexcelPUNIS-main\APU_CON_VAE_CONVERTIDO_v(153148).xlsx"

print(f"Abriendo: {xlsx_path}\n")

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

print("Verificando valores en columna I (CPC) y J (NP/EP/ND):")
print("=" * 80)

# Verificar filas de datos (después de encabezados)
test_rows = [13, 17, 18, 51, 55, 56, 57, 61, 62, 63, 93, 97, 98, 99, 132, 136, 137, 138, 142, 146]

for row in test_rows:
    cell_i = ws[f'I{row}']
    cell_j = ws[f'J{row}']
    
    value_i = cell_i.value
    value_j = cell_j.value
    type_i = cell_i.data_type
    type_j = cell_j.data_type
    
    status_i = "✓" if value_i else "❌"
    status_j = "✓" if value_j else "❌"
    
    print(f"Fila {row:4d}: I={status_i} '{value_i}' (type={type_i}) | J={status_j} '{value_j}' (type={type_j})")

wb.close()

print("\n\nSi las columnas I (CPC) están vacías, el problema es que los datos no se están extrayendo del PDF.")
