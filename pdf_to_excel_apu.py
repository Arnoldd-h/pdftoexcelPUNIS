"""
Script para convertir PDFs de Análisis de Precios Unitarios (APU) con VAE
a formato Excel estandarizado.

Autor: Automatizado
Fecha: Noviembre 2025
"""

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.packaging.core import DocumentProperties
from openpyxl.worksheet.datavalidation import DataValidation
import re
import os
import sys
import zipfile
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path


class APUConverter:
    """Clase para convertir PDFs de APU a Excel."""
    
    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.rubros = []
        self.header_info = {}
        
    def extract_header_info(self, text):
        """Extrae información del encabezado."""
        lines = text.split('\n')
        
        # Buscar nombre del profesional (primera línea generalmente)
        for line in lines[:5]:
            if 'ING.' in line.upper() or 'ARQ.' in line.upper() or 'LIC.' in line.upper():
                self.header_info['profesional'] = line.strip()
                break
        
        # Buscar proyecto
        for line in lines:
            if 'PROYECTO:' in line.upper():
                self.header_info['proyecto'] = line.strip()
                break
        
        # Buscar ubicación
        for line in lines:
            if 'UBICACION:' in line.upper():
                self.header_info['ubicacion'] = line.strip()
                break
                
    def parse_page(self, page):
        """Parsea una página del PDF y extrae los datos del rubro."""
        text = page.extract_text()
        if not text:
            return None
            
        # Extraer header info si no existe
        if not self.header_info:
            self.extract_header_info(text)
        
        rubro_data = {
            'numero_rubro': None,
            'unidad': None,
            'detalle': None,
            'cantidad': None,
            'hoja': None,
            'equipos': [],
            'mano_obra': [],
            'materiales': [],
            'transporte': [],
            'subtotal_m': 0,
            'subtotal_n': 0,
            'subtotal_o': 0,
            'subtotal_p': 0,
            'total_costo_directo': 0,
            'vae_total': 0,
            'indirectos_pct': 0,
            'indirectos_valor': 0,
            'utilidad_pct': 0,
            'utilidad_valor': 0,
            'costo_total': 0,
            'valor_unitario': 0,
            'texto_valor': '',
            'fecha': '',
        }
        
        lines = text.split('\n')
        
        # Buscar número de rubro, unidad, detalle
        for i, line in enumerate(lines):
            # Número de rubro
            match = re.search(r'RUBRO\s*:\s*(\d+)', line, re.IGNORECASE)
            if match:
                rubro_data['numero_rubro'] = int(match.group(1))
            
            # Unidad
            match = re.search(r'UNIDAD:\s*(\S+)', line, re.IGNORECASE)
            if match:
                rubro_data['unidad'] = match.group(1)
            
            # Detalle
            match = re.search(r'DETALLE\s*:\s*(.+)', line, re.IGNORECASE)
            if match:
                rubro_data['detalle'] = match.group(1).strip()
            
            # Hoja
            match = re.search(r'HOJA\s+(\d+)\s+DE\s+(\d+)', line, re.IGNORECASE)
            if match:
                rubro_data['hoja'] = f"HOJA {match.group(1)} DE {match.group(2)}"
                rubro_data['numero_pagina'] = int(match.group(1))
            
            # Fecha
            if 'LORETO,' in line.upper() or any(mes in line.upper() for mes in ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']):
                if 'DE 202' in line or 'DE 2024' in line or 'DE 2025' in line:
                    rubro_data['fecha'] = line.strip()
            
            # SON:
            if line.startswith('SON:'):
                rubro_data['texto_valor'] = line.strip()
            
            # ESPECIFICACIONES:
            if line.startswith('ESPECIFICACIONES:'):
                rubro_data['especificaciones'] = line.strip()
            
            # OBSERVACIONES:
            if line.startswith('OBSERVACIONES:'):
                rubro_data['observaciones'] = line.strip()
        
        # Extraer tablas
        tables = page.extract_tables()
        
        # Determinar sección actual
        current_section = None
        last_data_row = None  # Para capturar subtotales
        
        for table in tables:
            if not table:
                continue
                
            for row_idx, row in enumerate(table):
                if not row or all(cell is None for cell in row):
                    continue
                
                row_text = ' '.join([str(cell) if cell else '' for cell in row])
                first_cell = str(row[0]).upper() if row[0] else ''
                
                # Detectar cambio de sección
                if 'EQUIPO' in first_cell and 'DESCRIPCION' in first_cell:
                    current_section = 'equipo'
                    continue
                elif 'MANO DE OBRA' in first_cell and 'DESCRIPCION' in first_cell:
                    # Capturar subtotal M de la fila anterior (que tiene solo el valor)
                    if current_section == 'equipo' and last_data_row:
                        for cell in last_data_row:
                            if cell and self._is_number(cell):
                                rubro_data['subtotal_m'] = self._parse_number(cell)
                                break
                    current_section = 'mano_obra'
                    last_data_row = None
                    continue
                elif 'MATERIALES' in first_cell and 'DESCRIPCION' in first_cell:
                    # Capturar subtotal N de la fila anterior
                    if current_section == 'mano_obra' and last_data_row:
                        for cell in last_data_row:
                            if cell and self._is_number(cell):
                                rubro_data['subtotal_n'] = self._parse_number(cell)
                                break
                    current_section = 'materiales'
                    last_data_row = None
                    continue
                elif 'TRANSPORTE' in first_cell and 'DESCRIPCION' in first_cell:
                    # Capturar subtotal O de la fila anterior
                    if current_section == 'materiales' and last_data_row:
                        for cell in last_data_row:
                            if cell and self._is_number(cell):
                                rubro_data['subtotal_o'] = self._parse_number(cell)
                                break
                    current_section = 'transporte'
                    last_data_row = None
                    continue
                
                # Identificar filas de totales
                if 'TOTAL COSTO DIRECTO' in row_text:
                    # Capturar subtotal P de la fila anterior
                    if current_section == 'transporte' and last_data_row:
                        for cell in last_data_row:
                            if cell and self._is_number(cell):
                                rubro_data['subtotal_p'] = self._parse_number(cell)
                                break
                    
                    for cell in row:
                        if cell and self._is_number(cell):
                            val = self._parse_number(cell)
                            if val > 0:
                                rubro_data['total_costo_directo'] = val
                                break
                    current_section = None
                
                elif 'INDIRECTOS' in row_text:
                    # Extraer porcentaje del texto
                    match = re.search(r'(\d+\.?\d*)\s*%', row_text)
                    if match:
                        rubro_data['indirectos_pct'] = float(match.group(1)) / 100
                    # Extraer valor numérico
                    for cell in row[1:]:
                        if cell and self._is_number(cell):
                            rubro_data['indirectos_valor'] = self._parse_number(cell)
                            break
                
                elif 'UTILIDAD' in row_text and 'COSTO' not in row_text:
                    match = re.search(r'(\d+\.?\d*)\s*%', row_text)
                    if match:
                        rubro_data['utilidad_pct'] = float(match.group(1)) / 100
                    for cell in row[1:]:
                        if cell and self._is_number(cell):
                            rubro_data['utilidad_valor'] = self._parse_number(cell)
                            break
                
                elif 'COSTO TOTAL DEL RUBRO' in row_text:
                    for cell in row:
                        if cell and self._is_number(cell):
                            rubro_data['costo_total'] = self._parse_number(cell)
                            break
                
                elif 'VALOR UNITARIO' in row_text:
                    for cell in row:
                        if cell and self._is_number(cell):
                            rubro_data['valor_unitario'] = self._parse_number(cell)
                            break
                
                # Procesar filas de datos según la sección actual
                elif current_section:
                    # Si primera celda es None, puede ser fila de subtotal
                    if row[0] is None:
                        # Es probable que sea una fila de subtotal - guardar para siguiente sección
                        last_data_row = row
                    elif 'SUBTOTAL' not in first_cell:
                        row_data = self._extract_row_values_improved(row, current_section)
                        if row_data['descripcion'] and row_data['descripcion'].strip():
                            if current_section == 'equipo':
                                rubro_data['equipos'].append(row_data)
                            elif current_section == 'mano_obra':
                                rubro_data['mano_obra'].append(row_data)
                            elif current_section == 'materiales':
                                rubro_data['materiales'].append(row_data)
                            elif current_section == 'transporte':
                                rubro_data['transporte'].append(row_data)
                        last_data_row = row
        
        # Buscar cantidad en el texto (después del detalle)
        for line in lines:
            if rubro_data['detalle'] and rubro_data['detalle'] in line:
                continue
            # Buscar números solos, incluyendo formato con coma (1,058.84)
            numbers = re.findall(r'^\s*([\d,]+\.?\d*)\s*$', line)
            for num in numbers:
                # Convertir formato con coma a número
                val_str = num.replace(',', '')
                try:
                    val = float(val_str)
                    if val > 0 and val < 100000:
                        if rubro_data['cantidad'] is None:
                            rubro_data['cantidad'] = val
                except:
                    pass
        
        # Buscar VAE total desde el texto (está al final de la línea TOTAL COSTO DIRECTO)
        # Formato esperado: "TOTAL COSTO DIRECTO (M+N+O+P) 1.44 100.00% 97.08%"
        for line in lines:
            if 'TOTAL COSTO DIRECTO' in line.upper():
                # Buscar todos los porcentajes en la línea
                percentages = re.findall(r'(\d+\.?\d*)\s*%', line)
                if len(percentages) >= 2:
                    # El último porcentaje es el VAE total (ej: 97.08%)
                    # El primero suele ser 100.00%
                    rubro_data['vae_total'] = float(percentages[-1]) / 100
                elif len(percentages) == 1:
                    rubro_data['vae_total'] = float(percentages[0]) / 100
                break
        
        # Si no se encontró VAE total, calcularlo sumando los VAE de elementos
        if rubro_data['vae_total'] == 0:
            vae_sum = 0
            for items in [rubro_data['equipos'], rubro_data['mano_obra'], rubro_data['materiales'], rubro_data['transporte']]:
                for item in items:
                    if item.get('vae_elemento'):
                        vae_sum += item['vae_elemento']
            if vae_sum > 0:
                rubro_data['vae_total'] = vae_sum
        
        return rubro_data
    
    def _extract_row_values_improved(self, row, section):
        """Extrae los valores de una fila según la sección."""
        result = {
            'descripcion': '',
            'categoria': '',
            'cantidad': None,
            'tarifa': None,
            'costo_hora': None,
            'rendimiento': None,
            'costo': None,
            'peso_relativo': None,
            'cpc': '',
            'np_ep_nd': '',
            'vae_pct': None,
            'vae_elemento': None,
            'unidad': ''
        }
        
        if not row:
            return result
        
        # Convertir todas las celdas a string y limpiar
        cells = []
        for cell in row:
            if cell is not None:
                cells.append(str(cell).strip())
            else:
                cells.append('')
        
        # Procesar primera celda (descripción)
        if cells[0]:
            desc = cells[0]
            # Buscar si tiene código de categoría incluido (EO C1, EO D2, etc.)
            match = re.search(r'(.+?)\s+(EO\s*[A-Z]\d+)', desc)
            if match:
                result['descripcion'] = match.group(1).strip()
                result['categoria'] = match.group(2).strip()
            else:
                # Para Herramienta Menor con costo incluido en la descripción
                # Formato: "Herramienta Menor 5% de M.O. 0.07"
                match_hm = re.search(r'^(Herramienta\s+Menor.+?)\s+(\d+\.\d+)$', desc)
                if match_hm:
                    result['descripcion'] = match_hm.group(1).strip()
                    result['costo'] = float(match_hm.group(2))
                else:
                    result['descripcion'] = desc
        
        # Procesar según la sección
        if section == 'equipo':
            # EQUIPO - Según análisis del PDF, para Herramienta Menor:
            # Celda[0] = 'Herramienta Menor 5% de M.O. 0.07' (descripción con costo)
            # Celda[1-5] = None (vacíos)
            # Celda[6] = peso relativo (4.861%)
            # Celda[7] = CPC (4299217233)
            # Celda[8] = ND
            # Celda[9] = VAE % (40.00%)
            # Celda[10] = VAE Elemento (1.944%)
            if len(cells) >= 11:
                # Para equipos normales (con valores en columnas intermedias)
                if cells[1] and self._is_number(cells[1]):
                    result['cantidad'] = self._parse_number(cells[1])
                if cells[2] and self._is_number(cells[2]):
                    result['tarifa'] = self._parse_number(cells[2])
                if cells[3] and self._is_number(cells[3]):
                    result['costo_hora'] = self._parse_number(cells[3])
                if cells[4] and self._is_number(cells[4]):
                    result['rendimiento'] = self._parse_number(cells[4])
                # Solo sobrescribir costo si hay valor en celda[5] (para equipos normales)
                # Para Herramienta Menor, el costo ya se extrajo de la descripción
                if cells[5] and self._is_number(cells[5]):
                    result['costo'] = self._parse_number(cells[5])
                elif result['costo'] is None:
                    # Intentar obtener costo de celda[5] si existe
                    pass  # Ya se intentó arriba
                # Peso relativo, CPC, NP/EP/ND, VAE siempre en mismas posiciones
                if cells[6] and '%' in str(cells[6]):
                    result['peso_relativo'] = self._parse_percentage(cells[6])
                if cells[7] and re.match(r'^\d{9,12}$', str(cells[7])):
                    result['cpc'] = cells[7]
                if cells[8] and str(cells[8]).upper() in ['NP', 'EP', 'ND']:
                    result['np_ep_nd'] = str(cells[8]).upper()
                if cells[9] and '%' in str(cells[9]):
                    result['vae_pct'] = self._parse_percentage(cells[9])
                if cells[10] and '%' in str(cells[10]):
                    result['vae_elemento'] = self._parse_percentage(cells[10])
                    
        elif section == 'mano_obra':
            # MANO DE OBRA - Según análisis del PDF:
            # Celda[0] = descripción (ej: 'Maestro mayor ejec.obras civil EO C1')
            # Celda[1] = cantidad (0.01)
            # Celda[2] = jornal/hr (4.75)
            # Celda[3] = costo hora (0.05)
            # Celda[4] = rendimiento (0.3200)
            # Celda[5] = costo (0.02)
            # Celda[6] = peso relativo (1.389%)
            # Celda[7] = CPC (541210012)
            # Celda[8] = EP
            # Celda[9] = VAE % (100.00%)
            # Celda[10] = VAE Elemento (1.389%)
            if len(cells) >= 11:
                if cells[1] and self._is_number(cells[1]):
                    result['cantidad'] = self._parse_number(cells[1])
                if cells[2] and self._is_number(cells[2]):
                    result['tarifa'] = self._parse_number(cells[2])
                if cells[3] and self._is_number(cells[3]):
                    result['costo_hora'] = self._parse_number(cells[3])
                if cells[4] and self._is_number(cells[4]):
                    result['rendimiento'] = self._parse_number(cells[4])
                if cells[5] and self._is_number(cells[5]):
                    result['costo'] = self._parse_number(cells[5])
                if cells[6] and '%' in str(cells[6]):
                    result['peso_relativo'] = self._parse_percentage(cells[6])
                if cells[7] and re.match(r'^\d{9,12}$', str(cells[7])):
                    result['cpc'] = cells[7]
                if cells[8] and str(cells[8]).upper() in ['NP', 'EP', 'ND']:
                    result['np_ep_nd'] = str(cells[8]).upper()
                if cells[9] and '%' in str(cells[9]):
                    result['vae_pct'] = self._parse_percentage(cells[9])
                if cells[10] and '%' in str(cells[10]):
                    result['vae_elemento'] = self._parse_percentage(cells[10])
                    
        elif section in ['materiales', 'transporte']:
            # MATERIALES/TRANSPORTE: idx 0=desc, 1=vacio, 2=unidad, 3=cantidad, 4=precio, 5=costo, 6=peso, 7=cpc, 8=np_ep, 9=vae%, 10=vae_elem
            # Según análisis del PDF:
            # Celda[0] = descripción
            # Celda[1] = None (vacío)
            # Celda[2] = unidad (u, kg, rollo, etc.)
            # Celda[3] = cantidad (0.2000, 0.0130, etc.)
            # Celda[4] = precio unitario
            # Celda[5] = costo
            # Celda[6] = peso relativo (%)
            # Celda[7] = CPC
            # Celda[8] = NP/EP/ND
            # Celda[9] = VAE (%)
            # Celda[10] = VAE (%) Elemento
            if len(cells) >= 11:
                if cells[2] and cells[2] != 'None':
                    result['unidad'] = cells[2]
                if cells[3] and self._is_number(cells[3]):
                    result['cantidad'] = self._parse_number(cells[3])
                if cells[4] and self._is_number(cells[4]):
                    result['tarifa'] = self._parse_number(cells[4])
                if cells[5] and self._is_number(cells[5]):
                    result['costo'] = self._parse_number(cells[5])
                if cells[6] and '%' in str(cells[6]):
                    result['peso_relativo'] = self._parse_percentage(cells[6])
                if cells[7] and re.match(r'^\d{9,12}$', str(cells[7])):
                    result['cpc'] = cells[7]
                if cells[8] and str(cells[8]).upper() in ['NP', 'EP', 'ND']:
                    result['np_ep_nd'] = str(cells[8]).upper()
                if cells[9] and '%' in str(cells[9]):
                    result['vae_pct'] = self._parse_percentage(cells[9])
                if cells[10] and '%' in str(cells[10]):
                    result['vae_elemento'] = self._parse_percentage(cells[10])
        
        return result
    
    def _parse_data_row(self, row, rubro_data, table):
        """Método legacy - no se usa más, reemplazado por _extract_row_values_improved."""
        pass
    
    def _is_number(self, value):
        """Verifica si un valor es numérico."""
        if value is None:
            return False
        try:
            val_str = str(value).replace('%', '').replace(',', '.').strip()
            float(val_str)
            return True
        except (ValueError, TypeError):
            return False
    
    def _parse_number(self, value):
        """Convierte un valor a número."""
        if value is None:
            return 0
        try:
            val_str = str(value).replace('%', '').replace(',', '.').strip()
            return float(val_str)
        except (ValueError, TypeError):
            return 0
    
    def _parse_percentage(self, value):
        """Convierte un porcentaje a decimal."""
        if value is None:
            return 0
        try:
            val_str = str(value).strip()
            has_percent = '%' in val_str
            val_str = val_str.replace('%', '').replace(',', '.').strip()
            val = float(val_str)
            # Si venía con símbolo %, siempre dividir por 100
            if has_percent:
                return val / 100
            # Si no tiene %, asumimos que ya está en formato decimal
            return val
        except (ValueError, TypeError):
            return 0
    
    def extract_all_rubros(self):
        """Extrae todos los rubros del PDF."""
        with pdfplumber.open(self.pdf_path) as pdf:
            print(f"Procesando {len(pdf.pages)} páginas...")
            for i, page in enumerate(pdf.pages):
                print(f"  Procesando página {i+1}/{len(pdf.pages)}...", end='\r')
                rubro = self.parse_page(page)
                if rubro and rubro['numero_rubro']:
                    self.rubros.append(rubro)
            print(f"\n  Encontrados {len(self.rubros)} rubros.")
        return self.rubros
    
    def create_excel(self, output_path):
        """Crea el archivo Excel con el formato estandarizado exacto."""
        from openpyxl.styles import PatternFill, numbers
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ANALISIS DE PUNIS"
        
        # Formatos de número específicos para PUNIS
        NUMBER_FORMAT_2DEC = '###,##0.00'
        NUMBER_FORMAT_4DEC = '###,##0.0000'
        NUMBER_FORMAT_5DEC = '###,##0.00000'
        
        # Función auxiliar para escribir celda numérica con formato
        def write_numeric_cell(row, col, value, decimals=2, border=None):
            cell = ws.cell(row=row, column=col, value=value)
            if decimals == 4:
                cell.number_format = NUMBER_FORMAT_4DEC
            elif decimals == 5:
                cell.number_format = NUMBER_FORMAT_5DEC
            else:
                cell.number_format = NUMBER_FORMAT_2DEC
            if border:
                cell.border = border
            return cell
        
        # Configurar anchos de columna EXACTOS del original
        column_widths = {
            'A': 33.67, 'B': 8.67, 'C': 12.67, 'D': 14.55,
            'E': 14.67, 'F': 15.67, 'G': 13.0, 'H': 13.78,
            'I': 13.0, 'J': 10.78, 'K': 13.0, 'L': 12.78
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Estilos exactos del original
        font_title = Font(name='Cambria', size=13, bold=True)
        font_project = Font(name='Cambria', size=9, bold=True)
        font_header_section = Font(name='Cambria', size=12, bold=True)
        font_normal = Font(name='Cambria', size=10, bold=False)
        font_header_table = Font(name='Cambria', size=9, bold=True)
        font_total = Font(name='Cambria', size=10, bold=True)
        
        align_center_top = Alignment(horizontal='centerContinuous', vertical='top', wrap_text=True)
        align_justify_top = Alignment(horizontal='justify', vertical='top', wrap_text=True)
        align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        
        # Formatos de número EXACTOS del original PUNIS
        fmt_text = '@'  # Formato texto
        fmt_number = '###,##0.00'
        fmt_rendimiento = '###,##0.0000'
        fmt_peso_relativo = '0.000%'  # Porcentaje con 3 decimales (ej: 2.990%)
        fmt_vae_pct = '0.00%'  # Porcentaje con 2 decimales
        fmt_vae_elemento = '0.000%'  # Porcentaje con 3 decimales
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        border_left = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border_right = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border_middle = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        
        current_row = 0  # Empezamos desde 0, se incrementará a 1 al inicio
        total_rubros = len(self.rubros)
        
        for rubro_idx, rubro in enumerate(self.rubros, 1):
            start_row = current_row + 1
            
            # === FILA 1: Vacía ===
            current_row += 1
            ws.row_dimensions[current_row].height = 22.8
            
            # === FILA 2: Vacía ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            
            # === FILA 3: Nombre del profesional ===
            current_row += 1
            ws.row_dimensions[current_row].height = 49.95
            cell = ws.cell(row=current_row, column=1, value=self.header_info.get('profesional', '') + '\n')
            cell.font = font_title
            cell.alignment = align_center_top
            
            # === FILA 4: Vacía (fusionada A4:G4) ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws.cell(row=current_row, column=1, value='')
            
            # === FILA 5: Proyecto y Ubicación (fusionada A5:L5) ===
            current_row += 1
            ws.row_dimensions[current_row].height = 55.05
            proyecto_ubicacion = self.header_info.get('proyecto', '') + '\n' + self.header_info.get('ubicacion', '')
            ws.merge_cells(f'A{current_row}:L{current_row}')
            cell = ws.cell(row=current_row, column=1, value=proyecto_ubicacion)
            cell.font = font_project
            cell.alignment = align_justify_top
            
            # === FILA 6: Vacía (fusionada A6:G6) ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws.cell(row=current_row, column=1, value='')
            
            # === FILA 7: ANALISIS DE PRECIOS UNITARIOS + HOJA + DETERMINACION ===
            current_row += 1
            ws.row_dimensions[current_row].height = 16.95
            cell = ws.cell(row=current_row, column=1, value='                                   ANALISIS DE PRECIOS UNITARIOS')
            cell.font = font_header_section
            ws.cell(row=current_row, column=7, value=f'HOJA {rubro_idx} DE {total_rubros}')
            ws.cell(row=current_row, column=8, value='               DETERMINACION DEL VAE DEL RUBRO')
            
            # === FILA 8: RUBRO + UNIDAD ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value=f'RUBRO   :      {rubro["numero_rubro"]}').font = font_normal
            ws.cell(row=current_row, column=7, value=f'UNIDAD: {rubro["unidad"]}')
            
            # === FILA 9: DETALLE + CANTIDAD ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value=f'DETALLE :      {rubro["detalle"]}').font = font_normal
            if rubro['cantidad']:
                ws.cell(row=current_row, column=7, value=rubro['cantidad'])
            
            # === FILA 10: ESPECIFICACIONES/OBSERVACIONES o Número de página ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            if rubro.get('especificaciones'):
                ws.cell(row=current_row, column=1, value=rubro['especificaciones']).font = font_normal
                ws.cell(row=current_row, column=6, value=rubro.get('numero_pagina', rubro['numero_rubro']))
                ws.cell(row=current_row, column=7, value=4)
            elif rubro.get('observaciones'):
                ws.cell(row=current_row, column=1, value=rubro['observaciones']).font = font_normal
                ws.cell(row=current_row, column=6, value=rubro.get('numero_pagina', rubro['numero_rubro']))
                ws.cell(row=current_row, column=7, value=4)
            else:
                ws.cell(row=current_row, column=1, value='')
                ws.cell(row=current_row, column=6, value=rubro.get('numero_pagina', rubro['numero_rubro']))
                ws.cell(row=current_row, column=7, value=4)
            
            # === FILA 11: Vacía ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value='')
            
            # === FILA 12: ENCABEZADO EQUIPO ===
            current_row += 1
            ws.row_dimensions[current_row].height = 25.95
            headers_equipo = ['EQUIPO\nDESCRIPCION', '514704408', 'CANTIDAD\nA', 'TARIFA\nB', 
                            'COSTO HORA\nC=AxB', 'RENDIMIENTO\nR', 'COSTO\nD=CxR',
                            'Peso Relativo\nElemento (%)', 'CPC\nElemento', 'NP / EP /\nND',
                            'VAE (%)', 'VAE (%)\nElemento']
            for col, header in enumerate(headers_equipo, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = font_header_table
                cell.alignment = align_center_wrap
                cell.border = thin_border
            
            # === FILA 13: Datos de equipo ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            if rubro['equipos']:
                for equipo in rubro['equipos']:
                    cell_desc = ws.cell(row=current_row, column=1, value=equipo['descripcion'])
                    cell_desc.border = border_left
                    cell_desc.number_format = fmt_text
                    
                    # Escribir todos los campos del equipo con formatos correctos
                    # Usar 0 en lugar de None para valores numéricos vacíos
                    c = ws.cell(row=current_row, column=3, value=equipo.get('cantidad') if equipo.get('cantidad') is not None else 0)
                    c.number_format = fmt_number
                    
                    c = ws.cell(row=current_row, column=4, value=equipo.get('tarifa') if equipo.get('tarifa') is not None else 0)
                    c.number_format = fmt_number
                    
                    c = ws.cell(row=current_row, column=5, value=equipo.get('costo_hora') if equipo.get('costo_hora') is not None else 0)
                    c.number_format = fmt_number
                    
                    c = ws.cell(row=current_row, column=6, value=equipo.get('rendimiento') if equipo.get('rendimiento') is not None else 0)
                    c.number_format = fmt_rendimiento
                    
                    c = ws.cell(row=current_row, column=7, value=equipo.get('costo') if equipo.get('costo') is not None else 0)
                    c.number_format = fmt_number
                    
                    c = ws.cell(row=current_row, column=8, value=equipo.get('peso_relativo') if equipo.get('peso_relativo') is not None else 0)
                    c.number_format = fmt_peso_relativo
                    
                    # Asegurar que CPC sea string
                    cpc_val = str(equipo.get('cpc', '')) if equipo.get('cpc') is not None else ''
                    c = ws.cell(row=current_row, column=9, value=cpc_val)
                    c.number_format = fmt_text
                    
                    # Celda NP/EP/ND con formato texto explícito
                    c_vae = ws.cell(row=current_row, column=10, value=equipo.get('np_ep_nd', 'ND'))
                    c_vae.number_format = fmt_text
                    c_vae.data_type = 's'  # Forzar tipo string
                    
                    c = ws.cell(row=current_row, column=11, value=equipo.get('vae_pct', 0) if equipo.get('vae_pct') is not None else 0)
                    c.number_format = fmt_vae_pct
                    
                    c = ws.cell(row=current_row, column=12, value=equipo.get('vae_elemento') if equipo.get('vae_elemento') is not None else 0)
                    c.border = border_right
                    c.number_format = fmt_vae_elemento
                    
                    # Aplicar bordes a todas las celdas de la fila
                    for col in range(2, 12):
                        if col != 12:
                            ws.cell(row=current_row, column=col).border = border_middle
                    current_row += 1
                    ws.row_dimensions[current_row].height = 15.0
                current_row -= 1  # Compensar el incremento extra
            else:
                cell_desc = ws.cell(row=current_row, column=1, value='Herramienta Menor 5% de M.O.')
                cell_desc.border = border_left
                cell_desc.number_format = fmt_text
                c = ws.cell(row=current_row, column=7, value=rubro['subtotal_m'])
                c.number_format = fmt_number
                peso_rel = round(rubro['subtotal_m'] / rubro['total_costo_directo'], 5) if rubro['total_costo_directo'] > 0 else 0
                c = ws.cell(row=current_row, column=8, value=peso_rel)
                c.number_format = fmt_peso_relativo
                c = ws.cell(row=current_row, column=9, value='4299217233')
                c.number_format = fmt_text
                c_vae = ws.cell(row=current_row, column=10, value='ND')
                c_vae.number_format = fmt_text
                c_vae.data_type = 's'
                c = ws.cell(row=current_row, column=11, value=0.4)
                c.number_format = fmt_vae_pct
                c = ws.cell(row=current_row, column=12, value=round(peso_rel * 0.4, 5))
                c.border = border_right
                c.number_format = fmt_vae_elemento
                for col in range(2, 12):
                    ws.cell(row=current_row, column=col).border = border_middle
            
            # === FILA 14: SUBTOTAL M ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value='SUBTOTAL M').border = border_left
            ws.cell(row=current_row, column=7, value=rubro['subtotal_m'])
            for c in range(2, 12):
                ws.cell(row=current_row, column=c).border = border_middle
            ws.cell(row=current_row, column=12).border = border_right
            
            # === FILA 15: Vacía (altura pequeña) ===
            current_row += 1
            ws.row_dimensions[current_row].height = 4.95
            
            # === FILA 16: ENCABEZADO MANO DE OBRA ===
            current_row += 1
            ws.row_dimensions[current_row].height = 25.95
            headers_mo = ['MANO DE OBRA\nDESCRIPCION', '', 'CANTIDAD\nA', 'JORNAL/HR\nB',
                         'COSTO HORA\nC=AxB', 'RENDIMIENTO\nR', 'COSTO\nD=CxR',
                         'Peso Relativo\nElemento (%)', 'CPC\nElemento', 'NP / EP /\nND',
                         'VAE (%)', 'VAE (%)\nElemento']
            for col, header in enumerate(headers_mo, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = font_header_table
                cell.alignment = align_center_wrap
                cell.border = thin_border
            
            # === FILAS DE MANO DE OBRA ===
            for mo in rubro['mano_obra']:
                current_row += 1
                ws.row_dimensions[current_row].height = 15.0
                cell_desc = ws.cell(row=current_row, column=1, value=mo['descripcion'])
                cell_desc.border = border_left
                cell_desc.number_format = fmt_text
                ws.cell(row=current_row, column=2, value=mo.get('categoria', ''))
                c = ws.cell(row=current_row, column=3, value=mo.get('cantidad'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=4, value=mo.get('tarifa'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=5, value=mo.get('costo_hora'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=6, value=mo.get('rendimiento'))
                c.number_format = fmt_rendimiento
                c = ws.cell(row=current_row, column=7, value=mo.get('costo'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=8, value=mo.get('peso_relativo'))
                c.number_format = fmt_peso_relativo
                c = ws.cell(row=current_row, column=9, value=mo.get('cpc'))
                c.number_format = fmt_text
                c_vae = ws.cell(row=current_row, column=10, value=mo.get('np_ep_nd', 'EP'))
                c_vae.number_format = fmt_text
                c_vae.data_type = 's'
                c = ws.cell(row=current_row, column=11, value=mo.get('vae_pct', 1))
                c.number_format = fmt_vae_pct
                c = ws.cell(row=current_row, column=12, value=mo.get('vae_elemento'))
                c.border = border_right
                c.number_format = fmt_vae_elemento
                for col in range(2, 12):
                    ws.cell(row=current_row, column=col).border = border_middle
            
            # === SUBTOTAL N ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value='SUBTOTAL N').border = border_left
            ws.cell(row=current_row, column=7, value=rubro['subtotal_n'])
            for c in range(2, 12):
                ws.cell(row=current_row, column=c).border = border_middle
            ws.cell(row=current_row, column=12).border = border_right
            
            # === FILA VACÍA (altura pequeña) ===
            current_row += 1
            ws.row_dimensions[current_row].height = 4.95
            
            # === ENCABEZADO MATERIALES ===
            current_row += 1
            ws.row_dimensions[current_row].height = 25.95
            headers_mat = ['MATERIALES\nDESCRIPCION', '', '', 'UNIDAD\n', 'CANTIDAD\nA',
                          'PRECIO UNIT.\nB', 'COSTO\nC=AxB', 'Peso Relativo\nElemento (%)',
                          'CPC\nElemento', 'NP / EP /\nND', 'VAE (%)', 'VAE (%)\nElemento']
            for col, header in enumerate(headers_mat, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = font_header_table
                cell.alignment = align_center_wrap
                cell.border = thin_border
            
            # === FILAS DE MATERIALES ===
            for mat in rubro['materiales']:
                current_row += 1
                ws.row_dimensions[current_row].height = 15.0
                cell_desc = ws.cell(row=current_row, column=1, value=mat['descripcion'])
                cell_desc.border = border_left
                cell_desc.number_format = fmt_text
                ws.cell(row=current_row, column=4, value=mat.get('unidad', ''))
                c = ws.cell(row=current_row, column=5, value=mat.get('cantidad'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=6, value=mat.get('tarifa'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=7, value=mat.get('costo'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=8, value=mat.get('peso_relativo'))
                c.number_format = fmt_peso_relativo
                c = ws.cell(row=current_row, column=9, value=mat.get('cpc'))
                c.number_format = fmt_text
                c_vae = ws.cell(row=current_row, column=10, value=mat.get('np_ep_nd', 'EP'))
                c_vae.number_format = fmt_text
                c_vae.data_type = 's'
                c = ws.cell(row=current_row, column=11, value=mat.get('vae_pct', 1))
                c.number_format = fmt_vae_pct
                c = ws.cell(row=current_row, column=12, value=mat.get('vae_elemento'))
                c.border = border_right
                c.number_format = fmt_vae_elemento
                for col in range(2, 12):
                    ws.cell(row=current_row, column=col).border = border_middle
            
            # === SUBTOTAL O ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value='SUBTOTAL O').border = border_left
            ws.cell(row=current_row, column=7, value=rubro['subtotal_o'])
            for c in range(2, 12):
                ws.cell(row=current_row, column=c).border = border_middle
            ws.cell(row=current_row, column=12).border = border_right
            
            # === FILA VACÍA (altura pequeña) ===
            current_row += 1
            ws.row_dimensions[current_row].height = 4.95
            
            # === ENCABEZADO TRANSPORTE ===
            current_row += 1
            ws.row_dimensions[current_row].height = 25.95
            headers_trans = ['TRANSPORTE\nDESCRIPCION', '', '', 'UNIDAD\n', 'CANTIDAD\nA',
                            'TARIFA\nB', 'COSTO\nC=AxB', 'Peso Relativo\nElemento (%)',
                            'CPC\nElemento', 'NP / EP /\nND', 'VAE (%)', 'VAE (%)\nElemento']
            for col, header in enumerate(headers_trans, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = font_header_table
                cell.alignment = align_center_wrap
                cell.border = thin_border
            
            # === FILAS DE TRANSPORTE ===
            for trans in rubro['transporte']:
                current_row += 1
                ws.row_dimensions[current_row].height = 15.0
                cell_desc = ws.cell(row=current_row, column=1, value=trans['descripcion'])
                cell_desc.border = border_left
                cell_desc.number_format = fmt_text
                ws.cell(row=current_row, column=4, value=trans.get('unidad', ''))
                c = ws.cell(row=current_row, column=5, value=trans.get('cantidad'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=6, value=trans.get('tarifa'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=7, value=trans.get('costo'))
                c.number_format = fmt_number
                c = ws.cell(row=current_row, column=8, value=trans.get('peso_relativo'))
                c.number_format = fmt_peso_relativo
                c = ws.cell(row=current_row, column=9, value=trans.get('cpc'))
                c.number_format = fmt_text
                c_vae = ws.cell(row=current_row, column=10, value=trans.get('np_ep_nd', 'EP'))
                c_vae.number_format = fmt_text
                c_vae.data_type = 's'
                c = ws.cell(row=current_row, column=11, value=trans.get('vae_pct', 1))
                c.number_format = fmt_vae_pct
                c = ws.cell(row=current_row, column=12, value=trans.get('vae_elemento'))
                c.border = border_right
                c.number_format = fmt_vae_elemento
                for col in range(2, 12):
                    ws.cell(row=current_row, column=col).border = border_middle
            
            # === SUBTOTAL P ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value='SUBTOTAL P').border = border_left
            ws.cell(row=current_row, column=7, value=rubro['subtotal_p'])
            for c in range(2, 12):
                ws.cell(row=current_row, column=c).border = border_middle
            ws.cell(row=current_row, column=12).border = border_right
            
            # === FILA VACÍA ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            
            # === TOTAL COSTO DIRECTO ===
            current_row += 1
            ws.row_dimensions[current_row].height = 18.0
            ws.cell(row=current_row, column=3, value='514704408').border = thin_border
            cell = ws.cell(row=current_row, column=4, value='TOTAL COSTO DIRECTO (M+N+O+P)')
            cell.font = font_total
            cell.border = thin_border
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=6).border = thin_border
            c = ws.cell(row=current_row, column=7, value=rubro['total_costo_directo'])
            c.border = thin_border
            c.number_format = fmt_number
            # Columna H: 100% (peso relativo total)
            c = ws.cell(row=current_row, column=8, value=1)  # 100%
            c.border = thin_border
            c.number_format = fmt_peso_relativo  # Formato porcentaje 0.000%
            ws.cell(row=current_row, column=9).border = thin_border
            ws.cell(row=current_row, column=10).border = thin_border
            ws.cell(row=current_row, column=11).border = thin_border
            # Columna L: VAE Total del rubro
            c = ws.cell(row=current_row, column=12, value=rubro['vae_total'])
            c.border = thin_border
            c.number_format = fmt_vae_elemento  # Formato porcentaje 0.000%
            
            # === INDIRECTOS ===
            current_row += 1
            ws.row_dimensions[current_row].height = 18.0
            ws.cell(row=current_row, column=4, value='INDIRECTOS (%)').border = thin_border
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=6, value=rubro['indirectos_pct']).border = thin_border
            ws.cell(row=current_row, column=7, value=rubro['indirectos_valor']).border = thin_border
            
            # === UTILIDAD ===
            current_row += 1
            ws.row_dimensions[current_row].height = 18.0
            ws.cell(row=current_row, column=4, value='UTILIDAD (%)').border = thin_border
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=6, value=rubro['utilidad_pct']).border = thin_border
            ws.cell(row=current_row, column=7, value=rubro['utilidad_valor']).border = thin_border
            
            # === COSTO TOTAL DEL RUBRO ===
            current_row += 1
            ws.row_dimensions[current_row].height = 18.0
            ws.cell(row=current_row, column=4, value='COSTO TOTAL DEL RUBRO').border = thin_border
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=6).border = thin_border
            ws.cell(row=current_row, column=7, value=rubro['costo_total']).border = thin_border
            
            # === VALOR UNITARIO ===
            current_row += 1
            ws.row_dimensions[current_row].height = 21.0
            ws.cell(row=current_row, column=4, value='VALOR UNITARIO').border = thin_border
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=6).border = thin_border
            ws.cell(row=current_row, column=7, value=rubro['valor_unitario']).border = thin_border
            
            # === FILA VACÍA ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value='')
            
            # === SON: ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value=rubro['texto_valor'])
            
            # === ESTOS PRECIOS NO INCLUYEN IVA ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value='ESTOS PRECIOS NO INCLUYEN IVA')
            
            # === FILAS VACÍAS ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            
            # === FECHA ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            ws.cell(row=current_row, column=1, value=rubro['fecha'])
            
            # === FILA VACÍA AL FINAL DEL RUBRO (solo 1) ===
            current_row += 1
            ws.row_dimensions[current_row].height = 15.0
            
            # Agregar salto de página después de cada rubro
            from openpyxl.worksheet.pagebreak import Break
            ws.row_breaks.append(Break(id=current_row))
        
        # === CONFIGURACIÓN DE PÁGINA PARA IMPRESIÓN/PDF ===
        from openpyxl.worksheet.page import PageMargins
        
        # Configurar orientación y tamaño de papel
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.scale = 52  # Escala al 52%
        ws.page_setup.fitToPage = False
        
        # Configurar márgenes
        ws.page_margins = PageMargins(
            top=0.5,
            bottom=0.7,
            left=0.7,
            right=0.15,
            header=0.3,
            footer=0.3
        )
        
        # Configurar títulos de impresión (print titles)
        ws.print_title_rows = '1:1'
        
        # === PROPIEDADES DEL DOCUMENTO PARA PUNIS ===
        wb.properties.title = 'PUNIS'
        wb.properties.subject = 'Precios Unitarios'
        wb.properties.creator = 'PUNIS'
        wb.properties.category = self.header_info.get('profesional', '')
        
        # === AGREGAR VALIDACIÓN DE DATOS PARA COLUMNA J (NP/EP/ND) ===
        # Crear la validación de datos con la lista de opciones
        # showDropDown debe estar en False (no mostrar dropdown) pero la validación sigue activa
        dv = DataValidation(
            type="list",
            formula1='"NP,EP,ND"',
            allow_blank=False,
            showDropDown=False,  # No mostrar el dropdown pero validar
            showInputMessage=False,
            showErrorMessage=False
        )
        dv.error = 'El valor debe ser NP, EP o ND'
        dv.errorTitle = 'Entrada inválida'
        dv.prompt = 'Seleccione NP, EP o ND'
        dv.promptTitle = 'Tipo de origen'
        
        # Agregar la validación a la hoja
        ws.add_data_validation(dv)
        
        # Aplicar la validación a todas las celdas de columna J que contengan datos
        for row in range(1, ws.max_row + 1):
            cell_j = ws.cell(row=row, column=10)  # Columna J
            if cell_j.value in ['NP', 'EP', 'ND', 'NP / EP /\nND']:
                dv.add(cell_j)
        
        # === APLICAR FORMATO DE NÚMEROS A TODAS LAS CELDAS NUMÉRICAS ===
        # NOTA: Los formatos de porcentaje para columnas 8, 11, 12 ya se aplican
        #       durante la generación de cada sección (EQUIPO, M.O., etc.)
        # Columnas con formato ###,##0.00: C, D, E, G (columnas numéricas regulares)
        # Columna con formato ###,##0.0000: F (Rendimiento)
        # Columnas con formato porcentaje (YA APLICADO): H=8 (0.000%), K=11 (0.00%), L=12 (0.000%)
        for row in range(1, ws.max_row + 1):
            for col in [3, 4, 5, 7]:  # C, D, E, G - Solo columnas numéricas regulares
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '###,##0.00'
            # Columna F (6) - Rendimiento con 4 decimales
            cell_f = ws.cell(row=row, column=6)
            if cell_f.value is not None and isinstance(cell_f.value, (int, float)):
                cell_f.number_format = '###,##0.0000'
        
        # Guardar archivo
        wb.save(output_path)
        print(f"Archivo guardado: {output_path}")
        return output_path


def convert_to_shared_strings(input_path, output_path=None):
    """Convierte un archivo XLSX de inline strings a shared strings PRESERVANDO el orden."""
    
    if output_path is None:
        output_path = input_path
    
    print(f"Post-procesando para compatibilidad PUNIS (Shared Strings)...")
    
    # Crear directorio temporal
    temp_dir = input_path + '_temp'
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)
    
    try:
        # Extraer el XLSX
        with zipfile.ZipFile(input_path, 'r') as z:
            z.extractall(temp_dir)
        
        # Parsear el worksheet
        sheet_path = os.path.join(temp_dir, 'xl', 'worksheets', 'sheet1.xml')
        
        # Registrar namespace
        ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
        ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
        
        # Leer y parsear
        with open(sheet_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Recopilar todos los inline strings EN ORDEN DE APARICIÓN
        # EXCEPCIÓN: No convertir "NP", "EP", "ND" ya que PUNIS necesita que permanezcan inline
        shared_strings = []
        string_map = {}  # mapa de string -> índice
        vae_values = {'NP', 'EP', 'ND'}  # Valores que NO se convierten
        
        # Buscar todos los inline strings EN ORDEN
        # Patrón más flexible para capturar inline strings
        pattern = r'<c r="([^"]*)"([^>]*)t="inlineStr"([^>]*)><is><t>([^<]*)</t></is></c>'
        
        # Primera pasada: recopilar strings en orden de primera aparición
        for match in re.finditer(pattern, content):
            cell_ref = match.group(1)
            text = match.group(4)
            
            # NO convertir valores NP/EP/ND - deben permanecer inline
            if text in vae_values:
                continue
                
            if text not in string_map:
                string_map[text] = len(shared_strings)
                shared_strings.append(text)
        
        print(f"  Encontrados {len(shared_strings)} strings únicos (excluyendo NP/EP/ND)")
        
        # Reemplazar inline strings con referencias a shared strings
        # IMPORTANTE: NP/EP/ND permanecen como inline strings SIN MODIFICAR
        def replace_inline(match):
            full_match = match.group(0)
            cell_ref = match.group(1)
            attrs_before = match.group(2)
            attrs_after = match.group(3)
            text = match.group(4)
            
            # Si es un valor VAE (NP/EP/ND), mantener EXACTAMENTE como está
            if text in vae_values:
                return full_match
            
            # Para otros valores, convertir a shared string
            if text in string_map:
                idx = string_map[text]
                # Preservar atributos de estilo si existen
                style_match = re.search(r's="(\d+)"', attrs_before + attrs_after)
                if style_match:
                    style = style_match.group(1)
                    return f'<c r="{cell_ref}" s="{style}" t="s"><v>{idx}</v></c>'
                else:
                    return f'<c r="{cell_ref}" t="s"><v>{idx}</v></c>'
            else:
                # Por seguridad, mantener inline si no está en el mapa
                return full_match
        
        new_content = re.sub(pattern, replace_inline, content)
        
        # Escribir el worksheet modificado
        with open(sheet_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # Crear el archivo sharedStrings.xml
        shared_strings_path = os.path.join(temp_dir, 'xl', 'sharedStrings.xml')
        
        ss_content = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        ss_content += '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        ss_content += f'count="{len(shared_strings)}" uniqueCount="{len(shared_strings)}">'
        
        for s in shared_strings:
            # Escapar caracteres especiales
            s_escaped = s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            ss_content += f'<si><t>{s_escaped}</t></si>'
        
        ss_content += '</sst>'
        
        with open(shared_strings_path, 'w', encoding='utf-8') as f:
            f.write(ss_content)
        
        # Actualizar [Content_Types].xml para incluir sharedStrings
        content_types_path = os.path.join(temp_dir, '[Content_Types].xml')
        with open(content_types_path, 'r', encoding='utf-8') as f:
            ct_content = f.read()
        
        if 'sharedStrings' not in ct_content:
            # Agregar el override para sharedStrings
            insert_pos = ct_content.find('</Types>')
            override = '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
            ct_content = ct_content[:insert_pos] + override + ct_content[insert_pos:]
            
            with open(content_types_path, 'w', encoding='utf-8') as f:
                f.write(ct_content)
        
        # Actualizar xl/_rels/workbook.xml.rels para incluir relación con sharedStrings
        rels_path = os.path.join(temp_dir, 'xl', '_rels', 'workbook.xml.rels')
        with open(rels_path, 'r', encoding='utf-8') as f:
            rels_content = f.read()
        
        if 'sharedStrings' not in rels_content:
            # Encontrar el próximo rId
            rids = re.findall(r'rId(\d+)', rels_content)
            next_rid = max(int(r) for r in rids) + 1 if rids else 1
            
            insert_pos = rels_content.find('</Relationships>')
            rel = f'<Relationship Id="rId{next_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
            rels_content = rels_content[:insert_pos] + rel + rels_content[insert_pos:]
            
            with open(rels_path, 'w', encoding='utf-8') as f:
                f.write(rels_content)
        
        # Crear el nuevo XLSX
        if os.path.exists(output_path):
            os.remove(output_path)
        
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    z.write(file_path, arcname)
                    
        print(f"  Archivo final guardado: {output_path}")
        
    except Exception as e:
        print(f"Error en post-procesamiento: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # Limpiar
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            
    return output_path


def convert_pdf_to_excel(pdf_path, output_path=None):
    """
    Función principal para convertir un PDF de APU a Excel.
    
    Args:
        pdf_path: Ruta al archivo PDF
        output_path: Ruta de salida para el Excel (opcional)
    
    Returns:
        Ruta del archivo Excel generado
    """
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"No se encontró el archivo: {pdf_path}")
    
    if output_path is None:
        from datetime import datetime
        timestamp = datetime.now().strftime("%H%M%S")
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        output_path = os.path.join(os.path.dirname(pdf_path), f"{base_name}_CONVERTIDO_v({timestamp}).xlsx")
    
    print(f"Iniciando conversión de: {pdf_path}")
    
    converter = APUConverter(pdf_path)
    converter.extract_all_rubros()
    converter.create_excel(output_path)
    
    # Post-procesar para asegurar compatibilidad con PUNIS (Shared Strings)
    # La versión mejorada ahora preserva el orden y contenido correcto
    convert_to_shared_strings(output_path)
    
    return output_path


def main():
    """Función principal del script."""
    if len(sys.argv) < 2:
        # Si no se proporciona argumento, buscar PDFs en el directorio actual
        current_dir = os.path.dirname(os.path.abspath(__file__))
        pdf_files = list(Path(current_dir).glob("*.pdf"))
        
        if not pdf_files:
            print("Uso: python pdf_to_excel_apu.py <archivo.pdf> [archivo_salida.xlsx]")
            print("\nNo se encontraron archivos PDF en el directorio actual.")
            return
        
        print(f"Encontrados {len(pdf_files)} archivos PDF:")
        for i, pdf in enumerate(pdf_files, 1):
            print(f"  {i}. {pdf.name}")
        
        # Convertir el primer PDF encontrado
        pdf_path = str(pdf_files[0])
    else:
        pdf_path = sys.argv[1]
    
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        result = convert_pdf_to_excel(pdf_path, output_path)
        print(f"\n✓ Conversión completada exitosamente!")
        print(f"  Archivo generado: {result}")
    except Exception as e:
        print(f"\n✗ Error durante la conversión: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
