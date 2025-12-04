
from openpyxl import load_workbook

def inspect_rubro_5():
    print("Loading APU_CON_VAE_CONVERTIDO.xlsx...")
    wb = load_workbook('APU_CON_VAE_CONVERTIDO.xlsx')
    ws = wb.active
    
    print("Searching for Rubro 5...")
    start_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and "RUBRO   :      5" in str(cell_val):
            start_row = row
            print(f"Found Rubro 5 at row {start_row}")
            break
            
    if start_row:
        # Print the next 20 rows to cover the rubro content
        print(f"\n--- Data for Rubro 5 (Rows {start_row} to {start_row+20}) ---")
        print("Row | A | B | C | D | E | F | G | H | I | J | K | L")
        print("-" * 80)
        
        for r in range(start_row, start_row + 25):
            row_vals = []
            for c in range(1, 13): # Columns A to L
                cell = ws.cell(row=r, column=c)
                val = cell.value
                fmt = cell.number_format
                # Show value and format for critical columns
                if c in [8, 9, 10, 11, 12] and val is not None:
                    row_vals.append(f"{val} [{fmt}]")
                else:
                    row_vals.append(str(val))
            
            print(f"{r} | " + " | ".join(row_vals))

if __name__ == "__main__":
    inspect_rubro_5()
